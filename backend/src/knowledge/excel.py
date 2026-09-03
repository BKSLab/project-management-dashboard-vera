from __future__ import annotations

import re
import zipfile
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

UNNAMED_COLUMN = re.compile(r"Unnamed:\s*(\d+)")

Row = list[Any]
Sheet = tuple[str, list[Row]]


def extract_excel_text(content: bytes) -> str:
    """Извлекает видимые строки всех листов книги как «колонка: значение».

    Первая видимая строка листа считается шапкой; скрытые пользователем или
    свёрнутые группировкой строки пропускаются, а объединённые ячейки
    разворачиваются в каждую ячейку своего диапазона.

    Args:
        content: Бинарное содержимое книги .xlsx/.xlsm/.xls.

    Returns:
        Плоский текст книги, по одной записи на строку.

    Raises:
        ValueError: Если книга повреждена или её формат не поддерживается.
    """
    sheets = _read_sheets(content)
    rendered = [text for title, rows in sheets if (text := _render_sheet(title, rows))]
    return "\n".join(rendered).strip()


def _read_sheets(content: bytes) -> list[Sheet]:
    """Читает книгу движком по её реальному формату, а не по расширению."""
    try:
        return _read_xlsx(content)
    except (InvalidFileException, zipfile.BadZipFile, KeyError):
        pass
    try:
        return _read_xls(content)
    except xlrd.XLRDError as error:
        raise ValueError(f"Не удалось разобрать книгу Excel: {error}") from error


def _read_xlsx(content: bytes) -> list[Sheet]:
    """Читает .xlsx/.xlsm через openpyxl, разворачивая объединённые ячейки."""
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    try:
        sheets: list[Sheet] = []
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            _expand_merged_cells(worksheet)
            hidden = {
                index for index, dimension in worksheet.row_dimensions.items() if dimension.hidden
            }
            # min_row=1 задан явно: ключи row_dimensions — настоящие номера
            # строк Excel, и нумерация обхода обязана начинаться с первой.
            rows = [
                list(row)
                for index, row in enumerate(
                    worksheet.iter_rows(min_row=1, values_only=True), start=1
                )
                if index not in hidden
            ]
            sheets.append((worksheet.title, rows))
        return sheets
    finally:
        workbook.close()


def _expand_merged_cells(worksheet: Worksheet) -> None:
    """Копирует значение верхней левой ячейки во все ячейки объединения."""
    for merged_range in list(worksheet.merged_cells.ranges):
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        worksheet.unmerge_cells(str(merged_range))
        if value is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                worksheet.cell(row, column).value = value


def _read_xls(content: bytes) -> list[Sheet]:
    """Читает .xls через xlrd; объединённые ячейки восполняются протяжкой."""
    book = xlrd.open_workbook(file_contents=content, formatting_info=True)
    sheets: list[Sheet] = []
    for sheet in book.sheets():
        if sheet.visibility != 0:
            continue
        hidden = {index for index, info in sheet.rowinfo_map.items() if getattr(info, "hidden", 0)}
        rows = [
            [_from_xls_cell(cell, book.datemode) for cell in sheet.row(index)]
            for index in range(sheet.nrows)
            if index not in hidden
        ]
        sheets.append((sheet.name, _forward_fill(rows)))
    return sheets


def _from_xls_cell(cell: xlrd.sheet.Cell, datemode: int) -> Any:
    """Приводит ячейку xlrd к обычному типу Python."""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return None
    return cell.value


def _forward_fill(rows: list[Row]) -> list[Row]:
    """Протягивает значение предыдущей строки в пустые ячейки .xls-книги."""
    filled: list[Row] = []
    previous: Row = []
    for row in rows:
        current = list(row)
        for index, value in enumerate(current):
            if _is_blank(value) and index < len(previous):
                current[index] = previous[index]
        filled.append(current)
        previous = current
    return filled


def _render_sheet(title: str, rows: list[Row]) -> str:
    """Собирает текст одного листа: его название и записи строк под шапкой."""
    meaningful = [row for row in rows if any(not _is_blank(value) for value in row)]
    if not meaningful:
        return ""
    columns = _unique_columns(meaningful[0])
    records = [record for row in meaningful[1:] if (record := _render_row(columns, row))]
    if not records:
        return ""
    return "\n".join([f"Лист: {title}", *records])


def _unique_columns(header: Row) -> list[str]:
    """Строит уникальные имена колонок, подставляя «#N» вместо пустых."""
    columns: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(header):
        name = UNNAMED_COLUMN.sub(r"#\1", _to_text(value)) or f"#{index}"
        while name in used:
            name = f"{name}_dup"
        used.add(name)
        columns.append(name)
    return columns


def _render_row(columns: list[str], row: Row) -> str:
    """Склеивает непустые ячейки строки в «колонка: значение»."""
    parts: list[str] = []
    for index, value in enumerate(row):
        if _is_blank(value):
            continue
        column = columns[index] if index < len(columns) else f"#{index}"
        parts.append(f"{column}: {_to_text(value)}")
    return ", ".join(parts)


def _is_blank(value: Any) -> bool:
    """Проверяет, что ячейка пуста по смыслу, а не только по значению."""
    return value is None or (isinstance(value, str) and not value.strip())


def _to_text(value: Any) -> str:
    """Печатает значение ячейки без артефактов вроде «1.0» и «00:00:00»."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, datetime):
        return value.date().isoformat() if _is_midnight(value) else value.isoformat(sep=" ")
    if isinstance(value, date | time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_midnight(value: datetime) -> bool:
    """Определяет дату без значимого времени, чтобы не печатать «00:00:00»."""
    return (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0)
